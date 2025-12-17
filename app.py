from flask import Flask, render_template, request
from openpyxl import Workbook, load_workbook
import os

app = Flask(__name__)
excel_file = "squid_game_registrations.xlsx"

# Create Excel file with headers if not exists
if not os.path.exists(excel_file):
    wb = Workbook()
    ws = wb.active
    ws.title = "Registrations"
    ws.append(["Name", "Gender", "Birthdate", "Yearly Income"])
    wb.save(excel_file)

@app.route("/", methods=["GET", "POST"])
def register():
    if request.method == "POST":
        name = request.form["name"]
        gender = request.form["gender"]
        birthdate = request.form["birthdate"]
        income = request.form["income"]

        wb = load_workbook(excel_file)
        ws = wb.active
        ws.append([name, gender, birthdate, income])
        wb.save(excel_file)
        return "Registration Saved Successfully!"

    return render_template("index.html")

if __name__ == "__main__":
    app.run(debug=True)
